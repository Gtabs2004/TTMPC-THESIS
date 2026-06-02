"""One-off: dump structure of LoanLedgerEdited.xlsx."""
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent.parent.parent
XLSX = REPO_ROOT / "src" / "analytics" / "RISK Assesment" / "TTMPC_Credit_Risk" / "LoanLedgerEdited.xlsx"

if not XLSX.exists():
    print(f"missing: {XLSX}", file=sys.stderr)
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
print(f"Workbook: {XLSX.name}")
print(f"Sheets: {wb.sheetnames}\n")

print(f"Total sheets: {len(wb.sheetnames)}\n")

# Inspect first few sheets in detail.
for sheet_name in wb.sheetnames[:3]:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name!r} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    for i, r in enumerate(rows, start=1):
        print(f"  row {i}: {r}")
    print()

wb.close()
