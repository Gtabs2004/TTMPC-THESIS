"""
Generates TTMPC_Risk_Simulator.xlsx — a spreadsheet template that reproduces
the model's feature engineering and adds a rule-based heuristic risk score.

Run from the project root in the activated venv:
    python "src/analytics/RISK Assesment/build_simulator_template.py"

Output: TTMPC_Risk_Simulator.xlsx in the same folder as this script.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule


OUT_PATH = Path(__file__).resolve().parent / "TTMPC_Risk_Simulator.xlsx"


# ---------- Styling helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="1D6021")
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1D6021")
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
VALUE_FONT = Font(name="Calibri", size=11)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor="FFFBEA")
COMPUTED_FILL = PatternFill("solid", fgColor="EAF7EC")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER


def style_label(cell):
    cell.font = LABEL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER


def style_input(cell):
    cell.font = VALUE_FONT
    cell.fill = INPUT_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER


def style_computed(cell):
    cell.font = VALUE_FONT
    cell.fill = COMPUTED_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER


def auto_width(ws, cols_widths):
    for col, w in cols_widths.items():
        ws.column_dimensions[col].width = w


# ---------- Build workbook ----------
wb = Workbook()


# ===== Sheet 1: Input =====
ws1 = wb.active
ws1.title = "Input"
ws1["A1"] = "Loan Application Input"
ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color="1D6021")
ws1.merge_cells("A1:B1")

inputs = [
    ("Loan Amount (PHP)", 90000),
    ("Occupation",        "Cashier"),
    ("Annual Income (PHP) — leave blank if not declared", None),
    ("Advance Payments Count (0 for new applications)",   0),
]
for i, (label, val) in enumerate(inputs, start=2):
    ws1.cell(row=i, column=1, value=label)
    style_label(ws1.cell(row=i, column=1))
    c = ws1.cell(row=i, column=2, value=val)
    style_input(c)

ws1["A6"] = "Yellow cells are inputs — change them to test scenarios."
ws1["A6"].font = NOTE_FONT
ws1.merge_cells("A6:B6")

auto_width(ws1, {"A": 55, "B": 24})

# Named ranges (workbook-scoped)
def add_name(name, ref):
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn

add_name("LoanAmount",    "Input!$B$2")
add_name("Occupation",    "Input!$B$3")
add_name("AnnualIncome",  "Input!$B$4")
add_name("AdvPayCount",   "Input!$B$5")


# ===== Sheet 2: OccupationMap =====
ws2 = wb.create_sheet("OccupationMap")
headers2 = ["Occupation", "Tier", "Stability Score"]
for j, h in enumerate(headers2, start=1):
    c = ws2.cell(row=1, column=j, value=h)
    style_header(c)

OCC_ROWS = [
    ("Teaching", "Public_Sector_Institutional", 4),
    ("Retired Teacher", "Public_Sector_Institutional", 4),
    ("Government Employee", "Public_Sector_Institutional", 4),
    ("Adas Iii", "Public_Sector_Institutional", 4),
    ("Local Treasury Operations Offices Ii", "Public_Sector_Institutional", 4),
    ("Mpiw Employee", "Public_Sector_Institutional", 4),
    ("Admin Officer V", "Public_Sector_Institutional", 4),
    ("Adas Ii", "Public_Sector_Institutional", 4),
    ("Ada", "Public_Sector_Institutional", 4),
    ("Social Welfare Assistant", "Public_Sector_Institutional", 4),
    ("Senior Fire Inspector", "Public_Sector_Institutional", 4),
    ("Encoder", "Public_Sector_Institutional", 4),
    ("Assistant Pharmacist", "Public_Sector_Institutional", 4),
    ("Dietitian", "Public_Sector_Institutional", 4),
    ("Sb Member", "Public_Sector_Institutional", 4),
    ("Seafarer", "Private_Professional_Skilled", 3),
    ("Automotive Technician", "Private_Professional_Skilled", 3),
    ("Beautician", "Private_Professional_Skilled", 3),
    ("Caregiver", "Private_Professional_Skilled", 3),
    ("Assistant Embalmer", "Private_Professional_Skilled", 3),
    ("Cashier", "Service_Support", 2),
    ("Store Clerk", "Service_Support", 2),
    ("Security Guard", "Service_Support", 2),
    ("Asst. Station Manager", "Service_Support", 2),
    ("Barber", "Service_Support", 2),
    ("Business", "Entrepreneurial_Informal", 1),
    ("Vendor", "Entrepreneurial_Informal", 1),
    ("Entrepreneur", "Entrepreneurial_Informal", 1),
    ("Farmer", "Entrepreneurial_Informal", 1),
    ("Rice Dealer", "Entrepreneurial_Informal", 1),
    ("Small Business", "Entrepreneurial_Informal", 1),
    ("Unknown", "Unclassified_High_Risk", 2),
]
for i, row in enumerate(OCC_ROWS, start=2):
    for j, val in enumerate(row, start=1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = VALUE_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER

auto_width(ws2, {"A": 42, "B": 32, "C": 16})


# ===== Sheet 3: Features =====
ws3 = wb.create_sheet("Features")
ws3["A1"] = "Model Feature Engineering — exact same 5 features the Random Forest receives"
ws3["A1"].font = Font(name="Calibri", size=12, bold=True, color="1D6021")
ws3.merge_cells("A1:C1")

# Normalized Occupation helper (row 2)
ws3["A2"] = "Normalized Occupation (helper)"
ws3["B2"] = (
    '=IF(TRIM(Occupation)="","Unknown",'
    'IF(ISNUMBER(SEARCH("automative",Occupation)),"Automotive Technician",'
    'PROPER(TRIM(Occupation))))'
)
ws3["C2"] = "Title Case + handles blank + the 'Automative' typo"

# The 5 features
feature_rows = [
    ("Feature 1: LoanAmount",            "=LoanAmount",
        "Direct input. No transformation."),
    ("Feature 2: Stability_Score",       "=IFERROR(VLOOKUP(B2, OccupationMap!A:C, 3, FALSE), 2)",
        "Lookup tier score; unknown occupation falls back to 2 (Unclassified_High_Risk)."),
    ("Feature 3: Advance_Payment_Count", "=AdvPayCount",
        "Always 0 for new applications."),
    ("Feature 4: Income_Is_Missing",     '=IF(OR(AnnualIncome="", AnnualIncome=0), 1, 0)',
        "Binary flag. Missingness itself is a learned signal."),
    ("Feature 5: Repayment_Stress_Index",
        "=IF(B6=1, -999, (LoanAmount/12) / (AnnualIncome/12) * 100)",
        "Uses -999 sentinel when income is missing. DO NOT change this constant."),
]
for i, (label, formula, note) in enumerate(feature_rows, start=3):
    style_label(ws3.cell(row=i, column=1, value=label))
    style_computed(ws3.cell(row=i, column=2, value=formula))
    nc = ws3.cell(row=i, column=3, value=note)
    nc.font = NOTE_FONT
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    nc.border = BORDER

style_label(ws3.cell(row=2, column=1))
style_computed(ws3.cell(row=2, column=2))
nc = ws3.cell(row=2, column=3)
nc.font = NOTE_FONT
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
nc.border = BORDER

ws3["A9"] = (
    "These five cells (B3:B7) are exactly what the Python /api/risk/predict "
    "endpoint feeds to the trained Random Forest."
)
ws3["A9"].font = NOTE_FONT
ws3.merge_cells("A9:C9")

auto_width(ws3, {"A": 36, "B": 32, "C": 55})


# ===== Sheet 4: RiskHeuristic =====
ws4 = wb.create_sheet("RiskHeuristic")
ws4["A1"] = "Rule-based Heuristic — explainable proxy, NOT the actual Random Forest"
ws4["A1"].font = Font(name="Calibri", size=12, bold=True, color="B45309")
ws4.merge_cells("A1:C1")

ws4["A2"] = (
    "This sheet helps you demonstrate the directional logic in Excel "
    "without running the Python model. Use the live UI for the real probability."
)
ws4["A2"].font = NOTE_FONT
ws4.merge_cells("A2:C2")

heur_rows = [
    ("Loan Size Risk (0–40)",       "=MIN(40, Features!B3 / 5000)",
        "Each PHP 5,000 of principal adds 1 point, capped at 40."),
    ("Stability Risk (0–30)",       "=(4 - Features!B4) * 7.5",
        "Tier 4 = 0 pts, Tier 1 = 22.5 pts. Public sector is safest."),
    ("Income Missing Risk (0–20)",  "=IF(Features!B6=1, 20, 0)",
        "Missing income is itself a strong signal in the training data (~51% missing)."),
    ("Repayment Stress Risk (0–10)","=IF(Features!B6=1, 5, MIN(10, Features!B7 / 30))",
        "5 pts if income missing; otherwise scaled by stress index."),
]
for i, (label, formula, note) in enumerate(heur_rows, start=4):
    style_label(ws4.cell(row=i, column=1, value=label))
    style_computed(ws4.cell(row=i, column=2, value=formula))
    nc = ws4.cell(row=i, column=3, value=note)
    nc.font = NOTE_FONT
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    nc.border = BORDER

# Totals
style_label(ws4.cell(row=9,  column=1, value="Heuristic Risk Score (0–100)"))
style_computed(ws4.cell(row=9,  column=2, value="=B4+B5+B6+B7"))
style_label(ws4.cell(row=10, column=1, value="Heuristic Risk Class"))
style_computed(ws4.cell(row=10, column=2, value='=IF(B9>=50,"High Risk","Performing")'))
style_label(ws4.cell(row=11, column=1, value="Probability Proxy"))
pp = ws4.cell(row=11, column=2, value="=B9/100")
style_computed(pp)
pp.number_format = "0.0%"

ws4["A13"] = (
    "Caveat: the actual model is a Random Forest with ~100 decision trees. "
    "This heuristic captures the dominant directions but can disagree with the "
    "real model on edge cases. Always cite the API output as the official score."
)
ws4["A13"].font = NOTE_FONT
ws4["A13"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws4.merge_cells("A13:C13")
ws4.row_dimensions[13].height = 36

# Conditional formatting on B10 (risk class)
ws4.conditional_formatting.add(
    "B10",
    CellIsRule(operator="equal", formula=['"High Risk"'],
               fill=PatternFill("solid", fgColor="FCA5A5")),
)
ws4.conditional_formatting.add(
    "B10",
    CellIsRule(operator="equal", formula=['"Performing"'],
               fill=PatternFill("solid", fgColor="BBF7D0")),
)

auto_width(ws4, {"A": 32, "B": 22, "C": 55})


# ===== Sheet 5: Output =====
ws5 = wb.create_sheet("Output")
ws5["A1"] = "TTMPC Risk Assessment — Simulation Output"
ws5["A1"].font = Font(name="Calibri", size=14, bold=True, color="1D6021")
ws5.merge_cells("A1:B1")

out_rows = [
    ("Loan Amount",            "=LoanAmount",                  '"₱"#,##0.00'),
    ("Occupation (raw)",       "=Occupation",                  "@"),
    ("Annual Income",          '=IF(AnnualIncome="","(not declared)",AnnualIncome)', '"₱"#,##0.00'),
    ("",                       "",                             ""),
    ("Stability Score",        "=Features!B4",                 "0"),
    ("Income Missing Flag",    "=Features!B6",                 "0"),
    ("Repayment Stress Index", "=Features!B7",                 "0.00"),
    ("",                       "",                             ""),
    ("Heuristic Risk Class",   "=RiskHeuristic!B10",           "@"),
    ("Heuristic Probability",  "=RiskHeuristic!B11",           "0.0%"),
]
for i, (label, formula, fmt) in enumerate(out_rows, start=3):
    style_label(ws5.cell(row=i, column=1, value=label))
    c = ws5.cell(row=i, column=2, value=formula)
    style_computed(c)
    if fmt:
        c.number_format = fmt

ws5["A14"] = (
    "Reminder: this Output sheet uses the rule-based heuristic. "
    "The production probability comes from the Random Forest via /api/risk/predict."
)
ws5["A14"].font = NOTE_FONT
ws5["A14"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws5.merge_cells("A14:B14")
ws5.row_dimensions[14].height = 36

auto_width(ws5, {"A": 28, "B": 28})


# ===== Save =====
wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}")
